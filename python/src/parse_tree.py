import openpyxl
from itertools import product
import codecs


file_path= './output/parse-results.txt'
with open(file_path, 'w') as f:
    pass
codecs.open(file_path,mode ='w')

write_txt = codecs.open(file_path, encoding= 'utf-8' ,mode ='a')
read_txt = codecs.open('./src/sentences.txt',  encoding= 'utf-8',mode= 'r')

grammar_frame = openpyxl.load_workbook("./output/grammar.xlsx").active
# grammar rules start at row 17 and end at row 62
rules = {}
for line in range(17,62):
    rhs = grammar_frame.cell(row = line , column = 2).value
    rhs = rhs.replace('[+inv]','')
    rhs = rhs.replace('[-inv]','')
    lhs = grammar_frame.cell(row = line , column = 3).value
    lhs = tuple(lhs.split(' '))
    rules[lhs] = [rhs]
        

lexicon = {}
for line in range(1,15):
    lhs = grammar_frame.cell(row = line , column = 2).value
    rhs = grammar_frame.cell(row=line , column = 4).value
    attributes = [a.strip() for a in rhs.split(',')]
    lexicon[lhs] = attributes


def reverse_dict(dict):
    new_dict = {}
    for pos in dict:
        for word in dict[pos]:
            if word in new_dict:
                if isinstance(new_dict[word], list):
                    new_dict[word].append(pos)
                else:
                    new_dict[word] = [new_dict[word], pos]
            else:
                new_dict[word] = [pos]
    return new_dict

lexicon = reverse_dict(lexicon)

def translate_sentence(sentence, lexicon):
    words = sentence.lower().split()
    def word_definition(words):
        if len(words) > 1:
            single_word = words[0]
            s1 = []
            s2 = []
            if single_word in lexicon:
                s = word_definition(words[1:])
                if s:
                    s1 = lexicon[single_word]
                    s1 = [[a] for a in s1]
                    s1 = [x + y for x, y in product(s1, s)]
            if len(words) > 1:
                compound_word = str(words[0] + ' ' + words[1])
                if compound_word in lexicon:
                    s = word_definition(words[2:])
                    if not words[2:]: s =[[]]
                    if s:
                        s2 = lexicon[compound_word]
                        s2 = [[a] for a in s2]
                        s2 = [x + y for x, y in product(s2, s)]
            x = s1 + s2
            return x
        if len(words) == 1 and words[0] in lexicon:
            return [[a] for a in lexicon[words[0]]]
        return []
    return word_definition(words)

            
def get_rule_list(x,rules = rules):
    # return type [[rule],...]
    return_value =[]
    for rule in rules:
        if rule[0] == x:
            return_value.append(rule)
    return return_value

##########################################################


def current_agenda_check(current_agenda,rules = rules):
    new_agenda = []
    for i in range(0,len(current_agenda)):
        start , word , end ,rule_lst,history = current_agenda[i]
        new_history = [current_agenda[i]]
        for rule in rule_lst:
            if len(rule)==1:
                new_agenda+= [(start, rules[rule][0],end,get_rule_list(rules[rule][0]),new_history)]
            else:
                en = end
                key = False
                past_list = []
                for j in range(1,len(rule)):
                    key = False
                    for k in range(i,len(current_agenda)):
                        s,w,e,r,h = current_agenda[k]
                        if s == en and w == rule[j]:
                            en = e
                            key = True
                            past_list += [current_agenda[k]]
                            break
                if key:
                    new_agenda+= [(start,rules[rule][0],en,get_rule_list(rules[rule][0]),new_history + past_list)]
    return new_agenda

def past_agenda_check(current_agenda,past_agenda,rules = rules):
    new_agenda = []
    for i in range(0,len(past_agenda)):
        start,word,end, rule_lst,history = past_agenda[i]
        new_history = [past_agenda[i]]
        for rule in rule_lst:
            if len(rule)!=1:
                en =end 
                key= False
                past_list = []
                for j in range(1,len(rule)):
                    key = False
                    for s,w,e,r,h in current_agenda :
                        if s == en and w == rule[j]:
                            en = e
                            key = True
                            past_list += [(s,w,e,r,h)]
                            break
                if key:
                    new_agenda+= [(start,rules[rule][0],en, get_rule_list(rules[rule][0]) ,new_history + past_list)]
    for i in range(0,len(current_agenda)):
        start,word,end,rule_lst, history = current_agenda[i]
        new_history = [current_agenda[i]]
        for rule in rule_lst:
            if len(rule)!= 1:
                en =end 
                key= False
                past_list = []
                for j in range(1,len(rule)):
                    key = False
                    for s,w,e,r,h in past_agenda:
                        if s == en and w == rule[j]:
                            en = e
                            key = True
                            past_list += [(s,w,e,r,h)]
                            break
                if key:
                    new_agenda+= [(start,rules[rule][0],en, get_rule_list(rules[rule][0]) ,new_history + past_list)]
    return new_agenda

def bottom_up_parser(sentence):
    past_agenda = []
    current_agenda = []
    length_sentence = len(sentence)
    for c, word in enumerate(sentence,start=0):
        current_agenda.append((c,word,c+1,get_rule_list(word),[]))
        
    while True:
        new_agenda = []
        new_agenda += current_agenda_check(current_agenda)
        new_agenda += past_agenda_check(current_agenda,past_agenda)
        past_agenda+=current_agenda
        current_agenda = new_agenda
        for a in new_agenda:
            s,w,e,r,h = a
            if (s,w,e) == (0,'S',length_sentence):
                return a
        if not new_agenda:
                break
    return []
    
    
def print_tree(tree):
    if tree:
        samelevel = []
        newlevel = []
        for s,w,e,r,h in tree:
            samelevel += [s,w,e]
            newlevel+=h
        print_tree((newlevel))
        write_txt.write(str(samelevel) + '\n')
    

# sentence = ['HW', 'N', 'N', 'V', 'ADJ']
# a = bottom_up_parser(sentence)
# print_tree(a[3])



for c,line in enumerate(read_txt,start=1):
    write_txt.write(str(c) + ". "+line)
    sentences = set(tuple(sublist) for sublist in translate_sentence(line,lexicon))
    for sentence in sentences:
        s = [f"{index} {word}" for index, word in enumerate(sentence)]
        s = ' '.join(s)
        write_txt.write(str(s) + ' ' + str(len(sentence)) + '\n')
        a=bottom_up_parser(sentence)
        if not a:
            write_txt.write('Can not find a parser tree for this sentence.\n')
        else:
            s,w,e,r,h = a
            print_tree(h)
            write_txt.write(str([s,w,e])+'\n')
        

read_txt.close()
write_txt.close()
