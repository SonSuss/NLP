import openpyxl
from itertools import product
from functools import reduce
import os

class ReachTheEnd(Exception):
    pass

##test
# rules = {'S'  :[['NP','VP']],
#                'VP' :[['V'] , ['V' , 'NP'], ['V','PP']],
#                'NP' : [['ART','NP'] , ['N'], ['ADJ','NP'], ['NP','PREPS']],
#                'PREPS': [['PP'], ['PP' , 'PREPS'] , ['P' , 'NP']]}
# ends = ['V' , 'N' , 'P' ,'ART' , 'PP']


grammar_frame = openpyxl.load_workbook("./output/grammar.xlsx").active
# grammar rules start at row 17 and end at row 62
rules = {}
for line in range(17,62):
    lhs = grammar_frame.cell(row = line , column = 2).value
    lhs = lhs.replace('[+inv]','')
    lhs = lhs.replace('[-inv]','')
    rhs = grammar_frame.cell(row = line , column = 3).value
    attributes = [a for a in rhs.split(' ')]
    if lhs in rules:
        rules[lhs].append(attributes)
    else:
        rules[lhs] = [attributes]

# because of non terminal rules, we go through the width of the parser tree to generate the sentence
ends = []
for line in range(1,15):
    value = grammar_frame.cell(row = line, column = 2).value
    ends+=[value]


def apply_rules(s):
    return_value = []
    for node in s:
        if node in rules:
            applied_node = rules[node]
            return_value += [applied_node]
        else:
            return_value+= [[[node]]]
    return_value = list(product(*return_value))
    return_value = [reduce(lambda a,e: a + list(e) , x ,[]) for x in return_value]
    return return_value
        

def generate_sentences():
    sentences = []
    generated = []
    count = 0
    for rule in rules['S']:
        generated += [rule]
    while True:
        new_generated = []
        for sentence in generated:
            new_sentences = apply_rules(sentence)
            for sentence in new_sentences:
                if all(word in ends for word in sentence):
                    sentences.append(sentence)
                    new_sentences.remove(sentence)
                    count+=1
                    if count >= 10000:
                        print("Reach 10000 sentences")        
                        return sentences
            new_generated += new_sentences
        generated = new_generated


sentences = generate_sentences()

file_path='./output/sentences.txt'
with open(file_path, 'w') as f:
    pass
for sentence in sentences:
    with open(file_path, 'a') as f:
        f.write(' '.join(sentence) + "\n")

